import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from dataentry import get_amount,get_category,get_date,get_description
import matplotlib.pyplot as plt

class MyExcelfile:
    MyFile = "Monthly Expense.xlsx"
    Columns = ["date", "category", "amount", "description"]
    dateFORMAT = "%m-%d-%Y"

    @classmethod
    def initialize_excel(cls):
        try:
            pd.read_excel(cls.MyFile)
        except FileNotFoundError:
            df = pd.DataFrame(columns= cls.Columns)
            df.to_excel(cls.MyFile, header = True, index= False)

    @classmethod
    def add_entry(cls, date,  category, amount, description):
        new_entry = { "date" : date,
                     "category": category,
                     "amount": amount,
                     "description": description}
        df_new = pd.DataFrame([new_entry])

        try:
            # Load existing workbook
            book = load_workbook(cls.MyFile)

            # Check if sheet exists
            sheet_name = 'Sheet1'  # Adjust as per your sheet name
            with pd.ExcelWriter(cls.MyFile, engine='openpyxl',if_sheet_exists='overlay', mode='a') as writer:
                # Get existing data from sheet
                startrow =book[sheet_name].max_row
                df_new.to_excel(writer, sheet_name=sheet_name, index=False, header= False, startrow=startrow)

        except FileNotFoundError:
            print( "File doesn't Exist please call initialize_excel method again!! ")

        print("Entry done successfully!  ")  

    @classmethod
    def get_transaction(cls,sdate,edate):
        df = pd.read_excel(cls.MyFile)
        df["date"] = pd.to_datetime(df["date"], format = MyExcelfile.dateFORMAT)
        sdate = datetime.strptime(sdate,MyExcelfile.dateFORMAT)
        edate = datetime.strptime(edate,MyExcelfile.dateFORMAT)

        mask = (df["date"] >= sdate) & (df["date"]<= edate)
        filtered_df = df.loc[mask]

        if filtered_df.empty:
            print("There are no transactions done in the given date range.")
        else:
            print(f"Transactions from {sdate.strftime(MyExcelfile.dateFORMAT)} to {edate.strftime(MyExcelfile.dateFORMAT)}")
            print(filtered_df.to_string(index= False, formatters= {"date": lambda x:x.strftime(MyExcelfile.dateFORMAT)}))
        
        total_income = filtered_df[filtered_df["category"]== "Income"]["amount"].sum()
        total_expense = filtered_df[filtered_df["category"] == "Expense"]["amount"].sum()
        print("\n Summary: ")
        print(f"Total Income is $ {total_income:.2f}")
        print(f"Total Expense is $ {total_expense: .2f}")
        print(f"Net Savings is $ {(total_income - total_expense):.2f}")



def adddata():
    MyExcelfile.initialize_excel()
    date = get_date("Enter the date of Transaction in 'mm-dd-yyyy' format or press enter for today's date: ", allow_default= True)
    category = get_category()
    amount = get_amount()
    description = get_description()
    MyExcelfile.add_entry(date,category,amount,description)


def plot_transaction(df):
    df = pd.read_excel(MyExcelfile.MyFile, sheet_name='Sheet1')
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df.set_index('date', inplace = True) 

    income_df = (df[df["category"] == "Income"]).resample("D").sum().reindex(df.index , fill_value = 0)
    expense_df = (df[df["category"] == "Expense"]).resample("D").sum().reindex(df.index , fill_value = 0)

    plt.figure(figsize=(10,5))
    plt.plot(income_df.index, income_df["amount"],label = "Income" , color = 'g')
    plt.plot(expense_df.index, expense_df["amount"], label = "Expense" , color = 'r')
    plt.xlabel("Date")
    plt.ylabel("Amount")
    plt.title(" Income and Expense Summary Report ")
    plt.legend()
    plt.grid(True)
    plt.show()



def main():
    while True:
        print()
        print("\n1. Add a Transaction. ")
        print("2. View Transactions within the date range. ")
        print("3. Exit ")
        choice = input("Please choose an option from 1 to 3 : ")

        if choice == "1":
            adddata()
        elif choice == "2":
            sdate = get_date("Enter the start date of the Transaction in 'mm-dd-yyyy' format: ")
            edate = get_date("Enter the End date of the Transaction in 'mm-dd-yyyy' format: ")
            df = MyExcelfile.get_transaction(sdate,edate)
            plot_choice = input("Do you want to see the report in plot ? (y/n)   ").lower()
            if plot_choice == 'y':
                plot_transaction(df)
                
            elif plot_choice == 'n':
                break
            else:
                print("Invalid Entry: Please Choose 'y' for plot display or 'n' to return to main menu   ")

        elif choice == "3":
            print(" Exiting . . . ")
            break
        else:
            print("Invalid Option. Please choose 1 , 2 or 3 ")

if __name__ == "__main__":
    main()

